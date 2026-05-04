import io
import os
import re
import urllib.request
from urllib.parse import quote

import openpyxl
import pandas as pd
import requests

_IMAGE_CACHE = os.path.join(os.path.dirname(__file__), "..", "outputs", "image_urls.csv")

_ITEMS_COLS = ["Order ID", "Transaction ID", "Item Title", "Item Specs", "Quantity", "Item Price", "Thumbnail"]


def load_google_sheet(url: str) -> pd.DataFrame:
    sheet_id = _extract_sheet_id(url)
    html = _fetch_html(sheet_id)
    all_names = re.findall(r'docs-sheet-tab-caption">([^<]+)<', html)

    items_names = [n for n in all_names if n.startswith("Items -")]
    orders_names = [n for n in all_names if n.startswith("Orders -")]
    print(f"Found {len(items_names)} 'Items' sheets and {len(orders_names)} 'Orders' sheets.")

    items_df = _load_items(sheet_id, items_names)
    orders_df = _load_orders(sheet_id, orders_names)
    images_df = _load_image_urls(sheet_id, items_names)

    combined = items_df.merge(orders_df, on="Order ID", how="left")
    combined["Transaction ID"] = combined["Transaction ID"].astype(str).str.split(".").str[0]
    combined = combined.merge(images_df, on="Transaction ID", how="left")

    matched_stores = combined["store_name"].notna().sum()
    matched_images = combined["image_url"].notna().sum()
    print(f"Stores joined: {matched_stores}/{len(combined)} rows.")
    print(f"Images joined: {matched_images}/{len(combined)} rows.")

    return combined


def _load_items(sheet_id: str, names: list[str]) -> pd.DataFrame:
    frames = []
    for name in names:
        url = _csv_url(sheet_id, name)
        try:
            df = pd.read_csv(url)
            if not set(_ITEMS_COLS).issubset(set(df.columns)):
                df = pd.read_csv(url, header=None, names=_ITEMS_COLS)
            df["_sheet"] = name
            frames.append(df)
        except Exception as e:
            print(f"  Skipped '{name}': {e}")
    combined = pd.concat(frames, ignore_index=True)
    print(f"Loaded {len(combined)} item rows across {len(frames)} sheets.")
    return combined


def _load_orders(sheet_id: str, names: list[str]) -> pd.DataFrame:
    frames = []
    for name in names:
        url = _csv_url(sheet_id, name)
        try:
            df = pd.read_csv(url, usecols=["Order ID", "Store Name"])
            df = df.rename(columns={"Store Name": "store_name"})
            frames.append(df)
        except Exception as e:
            print(f"  Skipped '{name}': {e}")
    combined = pd.concat(frames, ignore_index=True).drop_duplicates(subset=["Order ID"])
    print(f"Loaded {len(combined)} unique orders across {len(frames)} sheets.")
    return combined


def _load_image_urls(sheet_id: str, items_names: list[str]) -> pd.DataFrame:
    """Return a DataFrame of Transaction ID → image_url, using a local cache if available."""
    cache_path = os.path.abspath(_IMAGE_CACHE)

    if os.path.exists(cache_path):
        print("Loading image URLs from cache...")
        return pd.read_csv(cache_path, dtype={"Transaction ID": str})

    print("Downloading XLSX to extract image URLs (one-time, will be cached)...")
    xlsx_url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=xlsx"
    r = requests.get(xlsx_url, allow_redirects=True, timeout=120)
    wb = openpyxl.load_workbook(io.BytesIO(r.content))

    records = []
    for name in items_names:
        if name not in wb.sheetnames:
            continue
        ws = wb[name]
        headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        if "Transaction ID" not in headers or "Thumbnail" not in headers:
            continue
        tid_idx = headers.index("Transaction ID")
        thumb_idx = headers.index("Thumbnail")
        for row in ws.iter_rows(min_row=2, values_only=False):
            tid = row[tid_idx].value
            thumb = row[thumb_idx].value
            if tid and thumb:
                # Extract URL from =IMAGE("url") formula
                match = re.search(r'https?://[^"\']+', str(thumb))
                if match:
                    records.append({"Transaction ID": str(int(float(tid))), "image_url": match.group(0)})

    df = pd.DataFrame(records).drop_duplicates(subset=["Transaction ID"])
    df.to_csv(cache_path, index=False)
    print(f"Cached {len(df)} image URLs to {cache_path}")
    return df


def _csv_url(sheet_id: str, sheet_name: str) -> str:
    return f"https://docs.google.com/spreadsheets/d/{sheet_id}/gviz/tq?tqx=out:csv&sheet={quote(sheet_name)}"


def _fetch_html(sheet_id: str) -> str:
    url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/edit"
    req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
    with urllib.request.urlopen(req) as r:
        return r.read().decode("utf-8")


def _extract_sheet_id(url: str) -> str:
    match = re.search(r"/spreadsheets/d/([a-zA-Z0-9_-]+)", url)
    if not match:
        raise ValueError("Invalid Google Sheets URL.")
    return match.group(1)
